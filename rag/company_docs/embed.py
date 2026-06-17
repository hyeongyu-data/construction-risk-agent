"""
사내 문서 (계약서/견적서/기성청구서) → 텍스트 추출 → 청킹 → 임베딩 → PostgreSQL(pgvector) 저장
지원 포맷: PDF, Excel (.xlsx), Word (.docx)

실행:
  python rag/company_docs/embed.py              # 전체 문서 (재)임베딩
  python rag/company_docs/embed.py --reset      # 테이블 초기화 후 전체 재임베딩
  python rag/company_docs/embed.py --file 송파아파트_원가견적서_2023.xlsx  # 단일 파일 추가
"""

import os
import re
import sys
import argparse
from pathlib import Path
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
load_dotenv(PROJECT_ROOT / ".env")

import psycopg2
from pgvector.psycopg2 import register_vector
from langchain_aws import BedrockEmbeddings

# ── 설정 ──────────────────────────────────────────────────────────
DOCS_DIR      = PROJECT_ROOT / "data" / "raw" / "company_docs"
EMBED_MODEL   = "amazon.titan-embed-text-v2:0"
COMPANY_ID    = "대성물산"
VECTOR_DIM    = 1024

CHUNK_SIZE    = 500
CHUNK_OVERLAP = 80
MIN_CHUNK_LEN = 50

DB_CONFIG = {
    "host":     os.getenv("DB_HOST", "localhost"),
    "port":     os.getenv("DB_PORT", "5432"),
    "dbname":   os.getenv("DB_NAME", "material_cost"),
    "user":     os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD", ""),
}


def get_connection():
    conn = psycopg2.connect(**DB_CONFIG)
    register_vector(conn)
    return conn


# ── 파일명 → 메타데이터 파싱 ──────────────────────────────────────
DOC_TYPE_MAP = {
    "공사도급계약서": "공사도급계약서",
    "하도급계약서":   "하도급계약서",
    "자재공급계약서": "자재공급계약서",
    "원가견적서":     "견적서",
    "견적서":         "견적서",
    "기성청구서":     "기성청구서",
}

def parse_meta(filename: str) -> dict:
    stem = Path(filename).stem
    year = (re.search(r'(20\d{2})', stem) or [None, ""])[1]
    doc_type = next((v for k, v in DOC_TYPE_MAP.items() if k in stem), "기타")
    project_id = stem.split("_")[0]
    return {"project_id": project_id, "doc_type": doc_type, "year": year}


# ── 포맷별 텍스트 추출 ────────────────────────────────────────────

def extract_pdf(path: Path) -> str:
    import pdfplumber
    parts = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            if txt.strip():
                parts.append(txt.strip())
            for table in (page.extract_tables() or []):
                rows = [" | ".join(str(c).replace("\n", " ") if c else "" for c in row)
                        for row in table]
                if rows:
                    parts.append("\n".join(rows))
    return "\n\n".join(parts)


def extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        rows = []
        for row in wb[sheet_name].iter_rows(values_only=True):
            cells = [str(v).strip() if v is not None else "" for v in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"[시트: {sheet_name}]\n" + "\n".join(rows))
    return "\n\n".join(parts)


def extract_docx(path: Path) -> str:
    from docx import Document
    from docx.oxml.ns import qn
    from docx.table import Table as DocxTable
    doc = Document(str(path))
    parts = []
    for element in doc.element.body:
        tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag
        if tag == "p":
            text = "".join(n.text or "" for n in element.iter() if n.tag == qn("w:t"))
            if text.strip():
                parts.append(text.strip())
        elif tag == "tbl":
            tbl = DocxTable(element, doc)
            for row in tbl.rows:
                cells = [c.text.strip().replace("\n", " ") for c in row.cells]
                if any(cells):
                    parts.append(" | ".join(cells))
    return "\n".join(parts)


EXTRACTORS = {"pdf": extract_pdf, "xlsx": extract_xlsx, "docx": extract_docx}


# ── 청킹 ─────────────────────────────────────────────────────────

def chunk_text(text: str) -> list[str]:
    sections = re.split(r'(?=제\d+조)', text)
    chunks = []
    for section in sections:
        section = section.strip()
        if not section:
            continue
        if len(section) <= CHUNK_SIZE:
            if len(section) >= MIN_CHUNK_LEN:
                chunks.append(section)
        else:
            start = 0
            while start < len(section):
                chunk = section[start:start + CHUNK_SIZE].strip()
                if len(chunk) >= MIN_CHUNK_LEN:
                    chunks.append(chunk)
                start += CHUNK_SIZE - CHUNK_OVERLAP
    return chunks


# ── DB 초기화 ─────────────────────────────────────────────────────

def init_table(conn, reset: bool = False):
    cur = conn.cursor()
    cur.execute("CREATE SCHEMA IF NOT EXISTS rag")
    if reset:
        cur.execute("DROP TABLE IF EXISTS rag.company_docs")
        print("[INFO] 테이블 초기화 완료")
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS rag.company_docs (
            id            TEXT PRIMARY KEY,
            company_id    TEXT,
            project_id    TEXT,
            doc_type      TEXT,
            year          TEXT,
            file_name     TEXT,
            file_format   TEXT,
            chunk_index   INTEGER,
            content       TEXT,
            embedding     vector({VECTOR_DIM})
        )
    """)
    conn.commit()
    cur.close()
    print("[INFO] rag.company_docs 테이블 준비 완료")


# ── 단일 파일 임베딩 ──────────────────────────────────────────────

def embed_file(file_path: Path, conn, embedder):
    fmt = file_path.suffix.lower().lstrip(".")
    if fmt not in EXTRACTORS:
        print(f"  [SKIP] 지원하지 않는 형식: {file_path.name}")
        return 0

    print(f"[처리] {file_path.name}")

    try:
        raw_text = EXTRACTORS[fmt](file_path)
    except Exception as e:
        print(f"  [ERROR] 추출 실패: {e}")
        return 0

    if not raw_text.strip():
        print(f"  [WARN] 텍스트 없음, 스킵")
        return 0

    chunks = chunk_text(raw_text)
    print(f"  청크 수: {len(chunks)}개")
    if not chunks:
        return 0

    meta = parse_meta(file_path.name)
    cur = conn.cursor()

    # 기존 데이터 삭제
    cur.execute("DELETE FROM rag.company_docs WHERE file_name = %s", (file_path.name,))
    deleted = cur.rowcount
    if deleted:
        print(f"  [INFO] 기존 {deleted}개 청크 삭제 후 재삽입")

    embeddings = embedder.embed_documents(chunks)

    for i, (chunk, emb) in enumerate(zip(chunks, embeddings)):
        chunk_id = f"{COMPANY_ID}_{file_path.stem}_{i}"
        cur.execute("""
            INSERT INTO rag.company_docs
                (id, company_id, project_id, doc_type, year, file_name, file_format, chunk_index, content, embedding)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (id) DO UPDATE SET
                content   = EXCLUDED.content,
                embedding = EXCLUDED.embedding
        """, (
            chunk_id, COMPANY_ID,
            meta["project_id"], meta["doc_type"], meta["year"],
            file_path.name, fmt, i, chunk, emb,
        ))

    conn.commit()
    cur.close()
    print(f"  → 저장 완료")
    return len(chunks)


# ── 메인 ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="사내 문서 임베딩")
    parser.add_argument("--reset", action="store_true", help="테이블 초기화 후 전체 재임베딩")
    parser.add_argument("--file",  type=str, default=None,
                        help="단일 파일만 추가/갱신 (파일명만 입력, 예: 송파아파트_원가견적서_2023.xlsx)")
    args = parser.parse_args()

    embedder = BedrockEmbeddings(
        model_id=EMBED_MODEL,
        region_name=os.getenv("AWS_BEDROCK_REGION", "us-east-1"),
    )
    conn = get_connection()
    init_table(conn, reset=args.reset)

    if args.file:
        file_path = DOCS_DIR / args.file
        if not file_path.exists():
            print(f"[ERROR] 파일 없음: {file_path}")
            conn.close()
            return
        embed_file(file_path, conn, embedder)
    else:
        files = sorted(
            f for f in DOCS_DIR.iterdir()
            if f.suffix.lower() in (".pdf", ".xlsx", ".docx") and not f.name.startswith("~")
        )
        if not files:
            print(f"[WARN] 처리할 파일 없음: {DOCS_DIR}")
            conn.close()
            return

        total = 0
        for f in files:
            total += embed_file(f, conn, embedder)

        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM rag.company_docs")
        total_in_db = cur.fetchone()[0]
        cur.close()
        print(f"\n[완료] 총 {total}개 청크 / DB 전체: {total_in_db}개")

    conn.close()


if __name__ == "__main__":
    main()
